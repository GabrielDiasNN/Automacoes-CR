# import-error e wrong-import-position: import de lib/python via sys.path.insert()
# dinamico abaixo, que o pylint nao resolve em tempo de analise estatica. Os demais
# limites de complexidade (too-many-locals/branches/statements) foram removidos deste
# disable amplo apos o refactor de process() em helpers dedicados (achado de revisao
# RB-01 G1); bare-except tambem removido apos tipar a excecao em formatar_excel (RB-01 Q1).
# pylint: disable=line-too-long, missing-class-docstring, consider-using-max-builtin, broad-exception-caught, import-error, wrong-import-position
# {
#   "name": "processar-receitas-bloqueadas",
#   "version": "2.3.2",
#   "skill": "python-oracle-migration",
#   "description": "Use when processing blocked recipes with state control, visual highlighting, and Excel generation."
# }
import json
import os
import sys
from datetime import datetime
from typing import Any

sys.path.insert(
    0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "lib", "python")
)
import oracledb
import pandas as pd
from automation_log import ensure_utf8_streams, make_logger
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from oracle_client import init_oracle_thick_mode
from oracle_extract import (
    OracleCredentials,
    compute_hash,
    fetch_all,
    resolve_oracle_credentials,
)
from oracle_retry import CircuitBreakerError, make_oracle_retry
from pydantic import BaseModel, Field, ValidationError

# Carregar ambiente (.env) do projeto raiz
load_dotenv(os.path.join(os.path.dirname(__file__), "..", ".env"))

ensure_utf8_streams()


log = make_logger("PY-PROCESS")


# --- CONTRATOS DE DADOS (Pydantic) ---
class RecipeRecord(BaseModel):
    key: str = Field(alias="Key")
    cor: str = Field(alias="Cor Rec.")
    ep: int = Field(alias="EP Rec.")
    pe: int = Field(alias="PE Rec")
    data_prod: str | None = Field(alias="Data Última Prod.", default="")
    data_bloqueio: str | None = Field(alias="Data Bloqueio", default="")


class StateFile(BaseModel):
    last_hash: str
    updated_at: str
    records: list[RecipeRecord]


# --- RESILIENCIA DE CONEXAO ---
_oracle_retry = make_oracle_retry()


@_oracle_retry
def fetch_data_with_retry(
    creds: OracleCredentials, sql_query: str, exec_id: str
) -> pd.DataFrame:
    log(
        "Estabelecendo conexao e extraindo dados (com Circuit Breaker)...",
        "INFO",
        exec_id,
    )
    columns, rows = fetch_all(creds, sql_query, exec_id, log)
    return pd.DataFrame(rows, columns=columns)


# too-many-locals: builder de template HTML com muitas variaveis de cor/texto por
# legibilidade (nomes descritivos em vez de indices/dicionarios); extrair helpers
# fragmentaria o template sem reduzir complexidade real. Nao coberto pelo refactor
# de RB-01 G1 (escopo daquele achado era process()); disable localizado aqui.
def gerar_html_artistico(  # pylint: disable=too-many-locals
    df_display: pd.DataFrame, stats: dict[str, int]
) -> str:
    # Cores Classicas
    cor_header = "#0f4c81"
    cor_header_text = "#ffffff"
    cor_sub_header = "#e8f1ff"
    cor_table_header = "#e5eef8"
    cor_table_border = "#cbd5e1"

    # Cores de Status
    cor_new_bg = "#f0fdf4"  # Verde muito claro para novos
    cor_mod_bg = "#fef9c3"  # Amarelo suave para alteradas
    cor_del_bg = "#f1f5f9"  # Cinza muito claro para liberadas (riscado)

    # Textos ASCII-Safe
    rel_title = "Painel Geral de Receitas Bloqueadas"
    ciclo_resumo = "Resumo do Ciclo:"
    col_cor = "Cor Rec."
    col_ult_prod = "Data Última Prod."
    col_bloqueio = "Data Bloqueio"
    col_status = "Status / Mudança"
    txt_intro = "Abaixo, a listagem completa de receitas sob bloqueio com os destaques do ciclo:"
    txt_footer_ob = "Favor consultar a planilha em anexo para o detalhamento técnico completo por Ordem de Beneficiamento (OB)."
    txt_footer_auto = "Mensagem automática."
    txt_warning = (
        "Por favor, atualizar os status com as previsões de liberar cada receita."
    )

    def formatar_numero_inteiro_html(valor: Any) -> str:
        """Renderiza numeros inteiros sem casas decimais no corpo do e-mail."""
        if pd.isna(valor):
            return ""
        if isinstance(valor, (bool, int)):
            return str(valor)
        if isinstance(valor, float) and valor.is_integer():
            return str(int(valor))
        try:
            valor_float = float(valor)
            if valor_float.is_integer():
                return str(int(valor_float))
        except (TypeError, ValueError):
            pass
        return str(valor)

    html = f"""
    <div style="font-family:'Segoe UI', Calibri, Arial, sans-serif; font-size:11pt; color:#1f2937; line-height:1.5;">
        <table role="presentation" cellspacing="0" cellpadding="0" border="0" width="100%" style="max-width:1100px; border-collapse:collapse; border:1px solid #d1d5db;">
            <thead>
                <tr>
                    <th colspan="6" style="background-color:{cor_header}; color:{cor_header_text}; padding:18px; text-align:center;">
                        <div style="font-size:17pt; font-weight:bold; margin-bottom:5px;">{rel_title}</div>
                        <div style="font-size:10.5pt; font-weight:normal; color:{cor_sub_header};">
                            Atualizado em {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}
                        </div>
                    </th>
                </tr>
            </thead>
            <tbody>
                <tr>
                    <td style="padding:16px;">
                        <div style="margin-bottom:20px; padding:12px; background-color:#f8fafc; border-left:4px solid {cor_header}; border-radius:4px;">
                            <b style="color:{cor_header}; font-size:12pt;">{ciclo_resumo}</b><br>
                            <table cellspacing="5" cellpadding="0" style="margin-top:5px; font-size:10.5pt;">
                                <tr>
                                    <td><span style="display:inline-block; width:12px; height:12px; background-color:{cor_new_bg}; border:1px solid #bbf7d0;"></span> Novas: <b>{stats['new']}</b></td>
                                    <td style="padding-left:20px;"><span style="display:inline-block; width:12px; height:12px; background-color:{cor_mod_bg}; border:1px solid #fef08a;"></span> Alteradas: <b>{stats['mod']}</b></td>
                                    <td style="padding-left:20px;"><span style="display:inline-block; width:12px; height:12px; background-color:{cor_del_bg}; border:1px solid {cor_table_border};"></span> Liberadas: <b>{stats['del']}</b></td>
                                </tr>
                            </table>
                        </div>
                        <p style="margin-bottom:10px;">{txt_intro}</p>
                        <table cellspacing="0" cellpadding="6" border="1" width="100%" style="border-collapse:collapse; border:1px solid {cor_table_border}; font-size:10pt;">
                            <thead>
                                <tr style="background-color:{cor_table_header};">
                                    <th style="border:1px solid {cor_table_border};">{col_cor}</th>
                                    <th style="border:1px solid {cor_table_border};">EP Rec.</th>
                                    <th style="border:1px solid {cor_table_border};">PE Rec</th>
                                    <th style="border:1px solid {cor_table_border};">{col_ult_prod}</th>
                                    <th style="border:1px solid {cor_table_border};">{col_bloqueio}</th>
                                    <th style="border:1px solid {cor_table_border};">{col_status}</th>
                                </tr>
                            </thead>
                            <tbody>
    """

    # Ordenar: Liberadas no fim, o resto por Cor/EP
    df_display["_sort_order"] = df_display["_change_type"].apply(
        lambda x: 1 if x == "DELETED" else 0
    )
    df_display = df_display.sort_values(by=["_sort_order", "Cor Rec.", "EP Rec."])

    for row in df_display.to_dict("records"):
        change_type = row["_change_type"]
        bg = "#ffffff"
        status_text = ""
        font_style = "normal"
        text_decor = "none"

        if change_type == "NEW":
            bg = cor_new_bg
            status_text = "✨ NOVO BLOQUEIO"
        elif change_type == "MODIFIED":
            bg = cor_mod_bg
            status_text = "⚠ DATA ALTERADA"
        elif change_type == "DELETED":
            bg = cor_del_bg
            status_text = "<b style='color:#059669;'>✅ LIBERADA</b>"
            font_style = "italic"
            text_decor = "line-through"
        else:
            status_text = "<span style='color:#94a3b8;'>BLOQUEIO MANTIDO</span>"

        html += f"""
                                <tr style="background-color:{bg}; font-style:{font_style}; text-decoration:{text_decor};">
                                    <td style="border:1px solid {cor_table_border}; text-align:center; font-weight:bold;">{row['Cor Rec.']}</td>
                                    <td style="border:1px solid {cor_table_border}; text-align:center;">{formatar_numero_inteiro_html(row['EP Rec.'])}</td>
                                    <td style="border:1px solid {cor_table_border}; text-align:center;">{formatar_numero_inteiro_html(row['PE Rec'])}</td>
                                    <td style="border:1px solid {cor_table_border}; text-align:center;">{row['Data Última Prod.']}</td>
                                    <td style="border:1px solid {cor_table_border}; text-align:center;">{row['Data Bloqueio']}</td>
                                    <td style="border:1px solid {cor_table_border}; text-align:center; font-size:8.5pt; font-weight:bold;">{status_text}</td>
                                </tr>
        """
    html += f"""
                            </tbody>
                        </table>
                        <p style="margin-top:15px; color:#b91c1c; font-weight:bold;">{txt_warning}</p>
                    </td>
                </tr>
            </tbody>
        </table>
        <p style="margin-top:20px; font-size:10pt; color:#64748b;">
            {txt_footer_ob}<br>
            <span style="font-size:8.5pt; color:#cbd5e1;">{txt_footer_auto}</span>
        </p>
    </div>
    """
    return html


# too-many-branches: laco de formatacao percorre celula a celula aplicando varias
# regras condicionais de estilo (fill, formato de data, borda); nao coberto pelo
# refactor de RB-01 G1 (escopo daquele achado era process()); disable localizado aqui.
def formatar_excel(file_path: str) -> None:  # pylint: disable=too-many-branches
    wb = load_workbook(file_path)
    header_fill = PatternFill(
        start_color="0F4C81", end_color="0F4C81", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True, size=11)
    status_err_fill = PatternFill(
        start_color="FECACA", end_color="FECACA", fill_type="solid"
    )
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
                if isinstance(cell.value, datetime):
                    cell.number_format = "DD/MM/YYYY HH:MM:SS"
                elif sheet_name == "OBsReceitasBloqueadas" and cell.column in [
                    11,
                    12,
                    15,
                    18,
                    19,
                    20,
                ]:
                    cell.number_format = "DD/MM/YYYY"
                if sheet_name == "OBsReceitasBloqueadas":
                    if ws.cell(row=int(cell.row or 0), column=22).value == "Divergente":
                        ws.cell(row=int(cell.row or 0), column=22).fill = (
                            status_err_fill
                        )
                    ws.cell(row=int(cell.row or 0), column=2).font = Font(bold=True)
        for col in ws.columns:
            max_length = 0
            column = str(getattr(col[0], "column_letter", "A"))
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except (TypeError, ValueError, AttributeError):
                    pass
            ws.column_dimensions[column].width = max_length + 2
    wb.save(file_path)


def _carregar_e_normalizar(
    creds: OracleCredentials, exec_id: str
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Le a SQL, busca no Oracle e retorna (df_raw, df_agreg) ja normalizados."""
    sql_path = os.path.join(os.path.dirname(__file__), "SQL-ReceitasBloqueadas.sql")
    with open(sql_path, encoding="utf-8") as f:
        sql_query = f.read()

    df_raw = fetch_data_with_retry(creds, sql_query, exec_id)
    df_raw = df_raw.rename(
        columns={
            "COR_REC": "Cor Rec.",
            "EP_REC": "EP Rec.",
            "PE_REC": "PE Rec",
            "DATA_ULT_PROD": "Data Última Prod.",
            "DATA_BLOQUEIO": "Data Bloqueio",
        }
    )
    # Falso positivo do analisador estatico apos fetch_all()+pd.DataFrame(...)
    # tornar o tipo de df_raw precisamente inferivel; a resolucao do overload
    # de .rename(columns=...) acima passa a inferir None daqui em diante. Com
    # pd.read_sql() (forma anterior) o retorno ja era nao inferivel e
    # mascarava essa mesma limitacao do analisador.
    # pylint: disable=unsupported-assignment-operation,unsubscriptable-object
    for col in ["Data Última Prod.", "Data Bloqueio"]:
        df_raw[col] = (
            pd.to_datetime(df_raw[col], errors="coerce")
            .dt.strftime("%d/%m/%Y")
            .fillna("")
        )

    cols_agreg = [
        "Cor Rec.",
        "EP Rec.",
        "PE Rec",
        "Data Última Prod.",
        "Data Bloqueio",
    ]
    df_agreg = (
        df_raw[cols_agreg].drop_duplicates().sort_values(by=["Cor Rec.", "EP Rec."])
    )
    # pylint: enable=unsupported-assignment-operation,unsubscriptable-object
    df_agreg["Key"] = (
        df_agreg["Cor Rec."].astype(str)
        + "_"
        + df_agreg["EP Rec."].astype(str)
        + "_"
        + df_agreg["PE Rec"].astype(str)
    )
    return df_raw, df_agreg


def _carregar_ultimo_state(state_path: str, exec_id: str) -> list[Any]:
    """Le e valida (Pydantic) o receitas_state.json anterior.

    Retorna a lista de records (vazia se o arquivo nao existir ou for invalido).
    """
    last_state_data: Any = {}
    if os.path.exists(state_path):
        try:
            with open(state_path, encoding="utf-8") as f:
                raw_json = json.load(f)
                # Validacao de Contrato de Dados (Pydantic)
                StateFile.model_validate(raw_json)
                last_state_data = raw_json
        except ValidationError as ve:
            log(f"Contrato de Dados Violado no Estado: {ve}", "ERROR", exec_id)
            last_state_data = {}
        except Exception as e:
            log(f"Erro ao carregar estado: {e}", "WARN", exec_id)
            last_state_data = {}

    records: list[Any] = (
        last_state_data.get("records", [])
        if isinstance(last_state_data, dict)
        else last_state_data
    )
    return records


def _computar_diff(
    df_agreg: pd.DataFrame, last_records: list[Any]
) -> tuple[dict[str, int], list[dict[Any, Any]]]:
    """Compara df_agreg com o ultimo estado conhecido; retorna (stats, diff_rows)."""
    df_last = pd.DataFrame(last_records)
    stats = {"new": 0, "mod": 0, "del": 0}
    diff_rows: list[dict[Any, Any]] = []

    if df_last.empty:
        df_diff_new = df_agreg.copy()
        df_diff_new["_change_type"] = "NEW"
        diff_rows = df_diff_new.to_dict("records")
        stats["new"] = len(df_diff_new)
        return stats, diff_rows

    df_merge = df_agreg.merge(
        df_last, on="Key", how="outer", indicator=True, suffixes=("", "_last")
    )

    new_mask = df_merge["_merge"] == "left_only"
    del_mask = df_merge["_merge"] == "right_only"

    mod_mask = (df_merge["_merge"] == "both") & (
        df_merge["Data Bloqueio"] != df_merge["Data Bloqueio_last"]
    )

    stats["new"] = int(new_mask.sum())
    stats["mod"] = int(mod_mask.sum())
    stats["del"] = int(del_mask.sum())

    if stats["new"] > 0:
        df_new = df_merge[new_mask][df_agreg.columns].copy()
        df_new["_change_type"] = "NEW"
        diff_rows.extend(df_new.to_dict("records"))

    if stats["mod"] > 0:
        df_mod = df_merge[mod_mask][df_agreg.columns].copy()
        df_mod["_change_type"] = "MODIFIED"
        diff_rows.extend(df_mod.to_dict("records"))

    if stats["del"] > 0:
        rename_cols = {col + "_last": col for col in df_agreg.columns if col != "Key"}
        df_deleted = (
            df_merge[del_mask][["Key"] + list(rename_cols.keys())]
            .rename(columns=rename_cols)
            .copy()
        )
        df_deleted["_change_type"] = "DELETED"
        diff_rows.extend(df_deleted.to_dict("records"))

    return stats, diff_rows


def _montar_df_display(
    df_agreg: pd.DataFrame, diff_rows: list[dict[Any, Any]]
) -> pd.DataFrame:
    """Monta o DataFrame de exibicao (STABLE/NEW/MODIFIED + linhas DELETED) para o HTML."""
    df_diff = pd.DataFrame(diff_rows)
    df_display = df_agreg.copy()
    df_display["_change_type"] = "STABLE"

    if not df_diff.empty:
        change_map = df_diff[
            df_diff["_change_type"].isin(["NEW", "MODIFIED"])
        ].set_index("Key")["_change_type"]
        if not change_map.empty:
            df_display["_change_type"] = (
                df_display["Key"].map(change_map).fillna("STABLE")
            )

        df_deleted = df_diff[df_diff["_change_type"] == "DELETED"]
        if not df_deleted.empty:
            df_display = pd.concat([df_display, df_deleted], ignore_index=True)

    return df_display


def _salvar_state_tmp(df_agreg: pd.DataFrame, state_path: str) -> None:
    """Grava o state.json.tmp com o snapshot atual de df_agreg (commit final via PowerShell).

    Nao usa write_state_tmp() de lib/python/oracle_extract.py aqui porque essa funcao
    persiste apenas {last_hash, updated_at}; este script precisa manter tambem "records"
    (a lista completa de receitas bloqueadas) no state, para computar o diff na proxima
    execucao — um schema de state estendido que a funcao compartilhada nao suporta.
    """
    records = df_agreg.to_dict(orient="records")
    state_data = {
        "last_hash": compute_hash(records),
        "updated_at": datetime.now().isoformat(),
        "records": records,
    }
    with open(state_path + ".tmp", "w", encoding="utf-8") as f:
        json.dump(state_data, f, ensure_ascii=False, indent=4)


def _gerar_saidas(
    df_agreg: pd.DataFrame,
    df_raw: pd.DataFrame,
    diff: tuple[dict[str, int], list[dict[Any, Any]]],
    html_path: str,
    exec_id: str,
) -> None:
    """Gera o Excel formatado e o HTML de notificacao a partir do diff computado.

    `diff` e o par (stats, diff_rows) retornado por _computar_diff(), agrupado num
    unico parametro para manter a assinatura desta funcao dentro do limite de
    argumentos do pre-commit (R0913/R0917).
    """
    stats, diff_rows = diff
    excel_path = os.path.join(os.path.dirname(__file__), "Receitas Bloqueadas.xlsx")
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df_agreg.drop(columns=["Key"], errors="ignore").to_excel(
            writer, sheet_name="ReceitasBloqueadas", index=False
        )
        df_raw.to_excel(writer, sheet_name="OBsReceitasBloqueadas", index=False)

    try:
        formatar_excel(excel_path)
        log("Excel formatado.", "INFO", exec_id)
    except Exception as e:
        log(f"Aviso Excel: {e}", "WARN", exec_id)

    df_display = _montar_df_display(df_agreg, diff_rows)
    html_content = gerar_html_artistico(df_display, stats)
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html_content)


def process() -> None:
    exec_id = sys.argv[1] if len(sys.argv) > 1 else "manual"
    log(
        "Iniciando processo V2 de Receitas Bloqueadas (Resiliente e Validado)",
        "INFO",
        exec_id,
    )

    html_path = os.path.join(os.path.dirname(__file__), "email_body.html")
    if os.path.exists(html_path):
        os.remove(html_path)

    creds = resolve_oracle_credentials(log, exec_id, require_tns_admin=True)

    def handle_exception(exc_type: Any, exc_value: Any, exc_traceback: Any) -> None:
        if issubclass(exc_type, KeyboardInterrupt):
            sys.__excepthook__(exc_type, exc_value, exc_traceback)
            return
        log(f"FALHA NAO TRATADA: {exc_value}", "ERROR", exec_id)
        sys.__excepthook__(exc_type, exc_value, exc_traceback)

    sys.excepthook = handle_exception

    if creds is None:
        log("Credenciais ou caminhos Oracle invalidos no ambiente.", "ERROR", exec_id)
        sys.exit(1)

    init_oracle_thick_mode(
        creds.client_lib,
        creds.tns_admin,
        lambda msg, lvl="INFO": log(msg, lvl, exec_id),
    )

    try:
        df_raw, df_agreg = _carregar_e_normalizar(creds, exec_id)

        state_path = os.path.join(os.path.dirname(__file__), "receitas_state.json")
        last_records = _carregar_ultimo_state(state_path, exec_id)
        stats, diff_rows = _computar_diff(df_agreg, last_records)

        state_tmp_path = state_path + ".tmp"
        if not diff_rows:
            if os.path.exists(state_tmp_path):
                log(
                    "Sem novas alteracoes, mas detectado estado temporario pendente de notificacao.",
                    "INFO",
                    exec_id,
                )
                sys.exit(0)
            log("Sem alteracoes relevantes detectadas (Idempotencia).", "INFO", exec_id)
            sys.exit(2)

        _salvar_state_tmp(df_agreg, state_path)
        _gerar_saidas(df_agreg, df_raw, (stats, diff_rows), html_path, exec_id)

        log("Processo concluido com alteracoes. Notificacao gerada.", "INFO", exec_id)
        sys.exit(0)

    except CircuitBreakerError:
        log("Circuit Breaker Aberto: Banco de dados inacessivel.", "ERROR", exec_id)
        sys.exit(1)
    except oracledb.DatabaseError as de:
        if de.args:
            error_obj = de.args[0]
            log(f"Erro DB: {error_obj.message}", "ERROR", exec_id)
        else:
            log(f"Erro DB desconhecido: {de}", "ERROR", exec_id)
        sys.exit(1)
    except Exception as e:
        log(f"Erro fatal: {str(e)}", "ERROR", exec_id)
        sys.exit(1)


if __name__ == "__main__":
    process()
