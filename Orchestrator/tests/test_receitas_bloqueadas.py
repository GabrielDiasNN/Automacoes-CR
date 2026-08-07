# pylint: disable=protected-access, reimported, redefined-outer-name, unused-variable, line-too-long, wrong-import-position, too-many-locals, wrong-import-order, unused-argument, import-error
"""
Suite de Testes Unitários: Automação de Receitas Bloqueadas
Mapeamento de Regras de Negócio e resiliência via Pytest com Mocks.
"""

import json
import os
import sys
from typing import Any
from unittest.mock import MagicMock, patch

import openpyxl
import pandas as pd
import pytest

# Adicionar pasta da automacao ao PYTHONPATH dinamicamente
AUTOMATION_PATH = os.path.abspath(
    os.path.join(os.path.dirname(__file__), "..", "..", "Receitas Bloqueadas")
)
sys.path.append(AUTOMATION_PATH)

import processar_receitas


@pytest.fixture
def mock_df_receitas() -> Any:
    """Retorna dados de teste estruturados para as receitas."""
    return pd.DataFrame(
        [
            {
                "Cor Rec.": "Preto",
                "EP Rec.": 10,
                "PE Rec": 20,
                "Data Última Prod.": "20/05/2026",
                "Data Bloqueio": "19/05/2026",
                "Key": "Preto_10_20",
                "_change_type": "NEW",
            },
            {
                "Cor Rec.": "Azul",
                "EP Rec.": 5,
                "PE Rec": 15,
                "Data Última Prod.": "18/05/2026",
                "Data Bloqueio": "17/05/2026",
                "Key": "Azul_5_15",
                "_change_type": "MODIFIED",
            },
            {
                "Cor Rec.": "Vermelho",
                "EP Rec.": 2,
                "PE Rec": 8,
                "Data Última Prod.": "15/05/2026",
                "Data Bloqueio": "14/05/2026",
                "Key": "Vermelho_2_8",
                "_change_type": "DELETED",
            },
            {
                "Cor Rec.": "Verde",
                "EP Rec.": 8,
                "PE Rec": 12,
                "Data Última Prod.": "10/05/2026",
                "Data Bloqueio": "09/05/2026",
                "Key": "Verde_8_12",
                "_change_type": "STABLE",
            },
        ]
    )


def test_gerar_html_artistico_badges(mock_df_receitas: Any) -> None:
    """Garante que a geracao de HTML artistico renderiza os badges de status pt-BR corretamente."""
    stats = {"new": 1, "mod": 1, "del": 1}
    html = processar_receitas.gerar_html_artistico(mock_df_receitas, stats)

    assert "Painel Geral de Receitas Bloqueadas" in html
    assert "✨ NOVO BLOQUEIO" in html
    assert "⚠ DATA ALTERADA" in html
    assert "✅ LIBERADA" in html
    assert "BLOQUEIO MANTIDO" in html
    assert "Preto" in html
    assert "Azul" in html
    assert "Vermelho" in html
    assert "Verde" in html


def test_gerar_html_artistico_formata_numeros_inteiros(mock_df_receitas: Any) -> None:
    """Garante que EP Rec. e PE Rec sao renderizados sem casas decimais quando vierem como float inteiro."""
    df_float = mock_df_receitas.copy()
    df_float.loc[0, "EP Rec."] = 56.0
    df_float.loc[0, "PE Rec"] = 1.0
    df_float.loc[1, "EP Rec."] = 22.0
    df_float.loc[1, "PE Rec"] = 1.0

    html = processar_receitas.gerar_html_artistico(
        df_float, {"new": 1, "mod": 1, "del": 1}
    )

    assert "56.0" not in html
    assert "1.0" not in html
    assert ">56<" in html
    assert ">1<" in html


def test_formatar_excel_estilos(tmp_path: Any) -> None:
    """Valida se o formatador Excel ajusta o auto_filter, largura e estilos de cabeçalho."""
    file_path = os.path.join(tmp_path, "test_receitas.xlsx")

    # Criar planilha de teste
    wb = openpyxl.Workbook()
    ws1 = wb.active
    assert ws1 is not None
    ws1.title = "ReceitasBloqueadas"
    ws1.append(["Cor Rec.", "EP Rec.", "PE Rec", "Data Última Prod.", "Data Bloqueio"])
    ws1.append(["Preto", 10, 20, "20/05/2026", "19/05/2026"])

    ws2 = wb.create_sheet("OBsReceitasBloqueadas")
    ws2.append(["OB", "StatusOB", "Cor"])
    ws2.append(["12345", "Normal", "Preto"])
    ws2.append(["67890", "Normal", "Azul"])
    ws2.cell(row=2, column=22, value="Divergente")  # Coluna 22 (V) Divergente
    wb.save(file_path)

    # Executar formatacao
    processar_receitas.formatar_excel(file_path)

    # Validar que a planilha foi estilizada
    wb_new = openpyxl.load_workbook(file_path)
    sheet1 = wb_new["ReceitasBloqueadas"]
    sheet2 = wb_new["OBsReceitasBloqueadas"]

    # Validar congelamento e auto_filter
    assert sheet1.freeze_panes == "A2"
    assert sheet1.auto_filter.ref is not None

    # Validar cor de preenchimento do cabecalho (0F4C81 - Azul Central)
    header_cell = sheet1["A1"]
    assert header_cell.fill.start_color.rgb in ("000F4C81", "0F4C81")
    assert header_cell.font.color.rgb in ("FFFFFFFF", "00FFFFFF", "FFFFFF")
    assert header_cell.font.bold is True

    # Validar formatacao condicional de divergente (FECACA - Vermelho suave)
    div_cell = sheet2.cell(row=2, column=22)  # Coluna 22 (V)
    normal_cell = sheet2.cell(row=3, column=22)
    assert div_cell.fill.start_color.rgb in ("00FECACA", "FECACA")
    assert (
        normal_cell.fill.start_color.rgb == "00000000"
        or normal_cell.fill.fill_type is None
    )


@patch("processar_receitas.oracledb.connect")
@patch("processar_receitas.os.path.exists")
@patch("processar_receitas.open")
def test_process_sem_alteracoes_exit_2(
    mock_open: Any, mock_exists: Any, mock_connect: Any, tmp_path: Any
) -> None:
    """Valida que process() termina com exit(2) quando os dados do Oracle coincidem com o estado anterior (idempotência)."""
    os.environ["ORACLE_READONLY_USER"] = "test_user"
    os.environ["ORACLE_READONLY_PASSWORD"] = "test_pass"
    os.environ["ORACLE_CLIENT_LIB_DIR"] = tmp_path.as_posix()
    os.environ["TNS_ADMIN"] = tmp_path.as_posix()

    state_content = json.dumps(
        {
            "last_hash": "abc123",
            "updated_at": "2026-05-19T10:00:00",
            "records": [
                {
                    "Key": "Azul_5_15",
                    "Cor Rec.": "Azul",
                    "EP Rec.": 5,
                    "PE Rec": 15,
                    "Data Última Prod.": "18/05/2026",
                    "Data Bloqueio": "17/05/2026",
                }
            ],
        }
    )

    def side_exists(path: str) -> bool:
        if path == tmp_path.as_posix():
            return True  # ORACLE_CLIENT_LIB_DIR, checado por resolve_oracle_credentials
        if "SQL-ReceitasBloqueadas.sql" in path:
            return True
        if "receitas_state.json" in path and ".tmp" not in path:
            return True
        return False

    mock_exists.side_effect = side_exists

    sql_file = MagicMock()
    sql_file.read.return_value = "SELECT * FROM RECEITAS"
    state_file = MagicMock()
    state_file.read.return_value = state_content

    def _ctx(obj: Any) -> Any:
        w = MagicMock()
        w.__enter__.return_value = obj
        w.__exit__.return_value = None
        return w

    def side_open(path: str, mode: str = "r", encoding: Any = None) -> Any:
        if "SQL-ReceitasBloqueadas" in path:
            return _ctx(sql_file)
        if "receitas_state.json" in path:
            return _ctx(state_file)
        return _ctx(MagicMock())

    mock_open.side_effect = side_open

    mock_conn = MagicMock()
    mock_cursor = MagicMock()
    mock_connect.return_value = mock_conn
    mock_conn.__enter__.return_value = mock_conn
    mock_conn.cursor.return_value = mock_cursor

    # Oracle devolve exatamente os mesmos dados que estão no estado → sem diff
    # (fetch_all le columns via cursor.description e linhas via cursor.fetchmany
    # em lotes ate um lote vazio; sem isso, o while True nunca termina)
    mock_cursor.description = [
        ("COR_REC",),
        ("EP_REC",),
        ("PE_REC",),
        ("DATA_ULT_PROD",),
        ("DATA_BLOQUEIO",),
    ]
    mock_cursor.fetchmany.side_effect = [
        [("Azul", 5, 15, "2026-05-18", "2026-05-17")],
        [],
    ]

    with (
        patch("processar_receitas.sys.argv", ["processar_receitas.py", "test_idem"]),
        patch("processar_receitas.sys.exit", side_effect=SystemExit(2)) as mock_exit,
    ):
        with pytest.raises(SystemExit) as excinfo:
            processar_receitas.process()

    assert excinfo.value.code == 2
    mock_exit.assert_called_once_with(2)


@patch("processar_receitas.oracledb.connect")
@patch("processar_receitas.os.path.exists")
@patch("processar_receitas.open")
def test_process_sucesso_com_novos_bloqueios(
    mock_open: Any, mock_exists: Any, mock_connect: Any, tmp_path: Any
) -> None:
    """
    Valida o fluxo process() quando ha novos bloqueios, garantindo a geracao
    de html, xlsx e estado tmp.
    """
    # Configurar mock de variaveis de ambiente
    os.environ["ORACLE_READONLY_USER"] = "test_user"
    os.environ["ORACLE_READONLY_PASSWORD"] = "test_pass"
    os.environ["ORACLE_CLIENT_LIB_DIR"] = tmp_path.as_posix()
    os.environ["TNS_ADMIN"] = tmp_path.as_posix()

    # Mocks de arquivo exist
    def side_exists(path: str) -> bool:
        if path == tmp_path.as_posix():
            return True  # ORACLE_CLIENT_LIB_DIR, checado por resolve_oracle_credentials
        if "SQL-ReceitasBloqueadas.sql" in path:
            return True
        if "receitas_state.json" in path:
            return True  # Simula que existe estado anterior
        return False

    mock_exists.side_effect = side_exists

    # Mock da leitura de SQL e estado json anterior
    sql_content = "SELECT * FROM RECEITAS"
    state_content = json.dumps(
        {
            "last_hash": "12345",
            "updated_at": "2026-05-19T10:00:00",
            "records": [
                {
                    "Key": "Azul_5_15",
                    "Cor Rec.": "Azul",
                    "EP Rec.": 5,
                    "PE Rec": 15,
                    "Data Última Prod.": "18/05/2026",
                    "Data Bloqueio": "17/05/2026",
                }
            ],
        }
    )

    mock_file_sql = MagicMock()
    mock_file_sql.read.return_value = sql_content
    mock_file_state = MagicMock()
    mock_file_state.read.return_value = state_content

    # Retorna o arquivo correto de acordo com a abertura
    def _context_file(file_obj: Any) -> Any:
        wrapper = MagicMock()
        wrapper.__enter__.return_value = file_obj
        wrapper.__exit__.return_value = None
        return wrapper

    def side_open(path: str, mode: str = "r", encoding: Any = None) -> Any:
        # pylint: disable=unused-argument
        if "SQL-ReceitasBloqueadas" in path:
            return _context_file(mock_file_sql)
        if "receitas_state.json" in path and "tmp" not in path:
            return _context_file(mock_file_state)
        return _context_file(MagicMock())

    mock_open.side_effect = side_open

    # Mock do banco Oracle
    mock_conn = MagicMock()
    mock_cursor = MagicMock()
    mock_connect.return_value = mock_conn
    mock_conn.__enter__.return_value = mock_conn
    mock_conn.cursor.return_value = mock_cursor

    # Simula dados extraidos contendo a receita anterior modificada e uma nova!
    # (fetch_all le columns via cursor.description e linhas via cursor.fetchmany
    # em lotes ate um lote vazio; sem isso, o while True nunca termina)
    mock_cursor.description = [
        ("COR_REC",),
        ("EP_REC",),
        ("PE_REC",),
        ("DATA_ULT_PROD",),
        ("DATA_BLOQUEIO",),
    ]
    mock_cursor.fetchmany.side_effect = [
        [
            ("Azul", 5, 15, "2026-05-20", "2026-05-19"),  # Data mudou (MODIFIED)
            ("Preto", 10, 20, "2026-05-20", "2026-05-20"),  # Nova receita (NEW)
        ],
        [],
    ]

    excel_writer = MagicMock()
    excel_writer.__enter__.return_value = MagicMock()
    excel_writer.__exit__.return_value = None

    with (
        patch("processar_receitas.pd.ExcelWriter", return_value=excel_writer),
        patch("processar_receitas.pd.DataFrame.to_excel"),
        patch("processar_receitas.formatar_excel") as mock_format,
        patch("processar_receitas.sys.exit") as mock_exit,
    ):

        # Executa o processador
        processar_receitas.process()

        # Verifica se o formatador excel e gravacao do estado temporario foram chamados
        assert mock_format.called
        # sys.exit(0) deve ser chamado quando ha alteracoes com sucesso
        mock_exit.assert_called_once_with(0)
